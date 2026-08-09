"""
rent.py — Live borough rent medians from the ONS PRMS dataset.

Strategy:
  1. Discover the latest downloadable PRMS workbook link from the ONS dataset page.
  2. Parse borough-level median rents from Table 2.7 ("All categories").
  3. Cache in-process for 24 hours.
  4. Return None on any failure so callers can fall back to static seed values.
"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone
from html import unescape
import io
import re
from typing import Optional

import httpx

ONS_DATASET_PAGE = (
    "https://www.ons.gov.uk/peoplepopulationandcommunity/housing/datasets/"
    "privaterentalmarketsummarystatisticsinengland"
)
DEFAULT_TTL_HOURS = 24

# London local authorities (32 boroughs + City of London)
LONDON_BOROUGHS = {
    "Barking and Dagenham",
    "Barnet",
    "Bexley",
    "Brent",
    "Bromley",
    "Camden",
    "City of London",
    "Croydon",
    "Ealing",
    "Enfield",
    "Greenwich",
    "Hackney",
    "Hammersmith and Fulham",
    "Haringey",
    "Harrow",
    "Havering",
    "Hillingdon",
    "Hounslow",
    "Islington",
    "Kensington and Chelsea",
    "Kingston upon Thames",
    "Lambeth",
    "Lewisham",
    "Merton",
    "Newham",
    "Redbridge",
    "Richmond upon Thames",
    "Southwark",
    "Sutton",
    "Tower Hamlets",
    "Waltham Forest",
    "Wandsworth",
    "Westminster",
}

_cache: dict[str, int] | None = None
_cache_at: datetime | None = None


def _normalise_borough(value: str) -> str:
    return " ".join(value.strip().split()).lower()


def _is_cache_valid(now: datetime) -> bool:
    if _cache is None or _cache_at is None:
        return False
    return (now - _cache_at) < timedelta(hours=DEFAULT_TTL_HOURS)


def _extract_latest_xlsx_url(dataset_html: str) -> str | None:
    pattern = re.compile(
        r'href="(?P<href>/file\?uri=/peoplepopulationandcommunity/housing/datasets/'
        r'privaterentalmarketsummarystatisticsinengland/[^"<>]+?\.xlsx)"',
        re.IGNORECASE,
    )
    match = pattern.search(dataset_html)
    if not match:
        return None
    href = unescape(match.group("href"))
    return f"https://www.ons.gov.uk{href}"


def _to_int(value: object) -> int | None:
    if isinstance(value, (int, float)):
        if value <= 0:
            return None
        return int(round(float(value)))
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned or cleaned in {".", "..", "-"}:
            return None
        try:
            parsed = float(cleaned)
        except ValueError:
            return None
        if parsed <= 0:
            return None
        return int(round(parsed))
    return None


def _parse_borough_rents_from_workbook(binary: bytes) -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    wb = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)

    table_name = None
    for sheet in wb.sheetnames:
        if sheet.replace(" ", "").lower().startswith("table2.7"):
            table_name = sheet
            break
    if table_name is None:
        return {}

    ws = wb[table_name]

    borough_lookup = {_normalise_borough(name): name for name in LONDON_BOROUGHS}
    rents: dict[str, int] = {}

    # Expected row shape: [None, LA Code, Area Code, Area, Count, Mean, LQ, Median, ...]
    for row in ws.iter_rows(min_row=8, values_only=True):
        area = row[3] if len(row) > 3 else None
        median = row[7] if len(row) > 7 else None
        if not isinstance(area, str):
            continue
        canonical = borough_lookup.get(_normalise_borough(area))
        if canonical is None:
            continue
        median_int = _to_int(median)
        if median_int is None:
            continue
        rents[canonical] = median_int

    return rents


async def fetch_borough_rents() -> Optional[dict[str, int]]:
    """
    Fetch latest available ONS borough rent medians for London.

    Returns
    -------
    dict[str, int] | None
        Mapping of borough -> median monthly rent (GBP), or None on failure.
    """
    global _cache, _cache_at

    now = datetime.now(timezone.utc)
    if _is_cache_valid(now):
        return _cache

    try:
        async with httpx.AsyncClient(timeout=15.0, headers={"User-Agent": "HouseApp/1.0"}) as client:
            page_resp = await client.get(ONS_DATASET_PAGE)
            page_resp.raise_for_status()

            xlsx_url = _extract_latest_xlsx_url(page_resp.text)
            if not xlsx_url:
                return None

            workbook_resp = await client.get(xlsx_url)
            workbook_resp.raise_for_status()

        parsed = _parse_borough_rents_from_workbook(workbook_resp.content)
        if not parsed:
            return None

        _cache = parsed
        _cache_at = now
        return parsed
    except Exception:
        return None
